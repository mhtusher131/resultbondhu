import pandas as pd
from io import BytesIO
from typing import List, Dict, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def parse_marks_from_excel(file_bytes: bytes) -> Tuple[List[Dict], List[str]]:
    """
    Parse marks from Excel file.
    Expected columns: Roll (or Roll No), Marks (or Total Marks)
    Returns (list of {roll, marks}, list of errors)
    """
    errors = []
    records = []

    try:
        df = pd.read_excel(BytesIO(file_bytes))
    except Exception as e:
        return [], [f"Could not read Excel file: {str(e)}"]

    # Normalize column names
    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]

    # Find roll column
    roll_col = next((c for c in df.columns if "roll" in c), None)
    marks_col = next((c for c in df.columns if "marks" in c or "total" in c or "obtained" in c), None)

    if not roll_col:
        return [], ["Could not find 'Roll' column. Columns found: " + ", ".join(df.columns)]
    if not marks_col:
        return [], ["Could not find 'Marks' column. Columns found: " + ", ".join(df.columns)]

    for idx, row in df.iterrows():
        try:
            roll = int(row[roll_col])
            marks_val = row[marks_col]
            if pd.isna(marks_val):
                errors.append(f"Row {idx+2}: Roll {roll} — marks is empty, skipped.")
                continue
            marks = float(marks_val)
            if marks < 0 or marks > 100:
                errors.append(f"Row {idx+2}: Roll {roll} — marks {marks} out of range (0-100).")
                continue
            records.append({"roll": roll, "marks": marks})
        except (ValueError, TypeError) as e:
            errors.append(f"Row {idx+2}: Parse error — {str(e)}")

    return records, errors


def parse_marks_from_csv(file_bytes: bytes) -> Tuple[List[Dict], List[str]]:
    """Parse marks from CSV file."""
    errors = []
    records = []

    try:
        df = pd.read_csv(BytesIO(file_bytes))
    except Exception as e:
        return [], [f"Could not read CSV file: {str(e)}"]

    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
    roll_col = next((c for c in df.columns if "roll" in c), None)
    marks_col = next((c for c in df.columns if "marks" in c or "total" in c), None)

    if not roll_col or not marks_col:
        return [], [f"Required columns missing. Found: {list(df.columns)}"]

    for idx, row in df.iterrows():
        try:
            roll = int(row[roll_col])
            marks = float(row[marks_col])
            if 0 <= marks <= 100:
                records.append({"roll": roll, "marks": marks})
            else:
                errors.append(f"Row {idx+2}: marks {marks} out of range")
        except Exception as e:
            errors.append(f"Row {idx+2}: {str(e)}")

    return records, errors


def generate_marks_template_excel(subject_name: str, subject_code: str, students: list) -> bytes:
    """Generate a teacher marks-input template Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{subject_name} Marks"

    # Header info
    ws.merge_cells("A1:D1")
    ws["A1"] = f"ResultBondhu — Mark Entry Template"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1D9E75")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Subject: {subject_name} (Code: {subject_code})"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Column headers
    headers = ["Roll No", "Student Name", "Marks (0-100)", "Remarks"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="085041")
        cell.alignment = Alignment(horizontal="center")

    # Student rows
    for row_idx, student in enumerate(students, 5):
        ws.cell(row=row_idx, column=1, value=student["roll"])
        ws.cell(row=row_idx, column=2, value=student["full_name"])
        ws.cell(row=row_idx, column=3, value="")  # teacher fills this
        ws.cell(row=row_idx, column=4, value="")

        # Style alternating rows
        fill_color = "E1F5EE" if row_idx % 2 == 0 else "FFFFFF"
        for col in range(1, 5):
            ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor=fill_color)

        # Add border
        thin = Side(style="thin", color="CCCCCC")
        for col in range(1, 5):
            ws.cell(row=row_idx, column=col).border = Border(
                left=thin, right=thin, top=thin, bottom=thin
            )

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 20

    # Instructions
    last_row = 4 + len(students) + 2
    ws.cell(row=last_row, column=1, value="Instructions:").font = Font(bold=True, color="085041")
    ws.cell(row=last_row + 1, column=1, value="• Enter marks in column C only (0 to 100)")
    ws.cell(row=last_row + 2, column=1, value="• Do not change Roll No or Name columns")
    ws.cell(row=last_row + 3, column=1, value="• Save and send back to the exam controller")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_result_excel(summary) -> bytes:
    """Generate full result sheet as Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grade Sheet"

    # Title
    ws.merge_cells("A1:J1")
    ws["A1"] = summary.exam.college.name.upper()
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1D9E75")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = f"HSC {summary.exam.year} — {summary.exam.term} | {summary.exam.group} Group"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Stats row
    stats_row = 4
    stats = [
        ("Appeared", summary.appeared),
        ("Passed", summary.passed),
        ("Failed", summary.failed),
        ("Pass Rate", f"{summary.pass_rate:.1f}%"),
        ("Avg GPA", summary.average_gpa),
        ("Highest GPA", summary.highest_gpa),
        ("Lowest GPA", summary.lowest_gpa),
        ("A+ Count", summary.a_plus_count),
    ]
    for col, (label, val) in enumerate(stats, 1):
        label_cell = ws.cell(row=stats_row, column=col, value=label)
        label_cell.font = Font(bold=True, color="FFFFFF", size=9)
        label_cell.fill = PatternFill("solid", fgColor="085041")
        label_cell.alignment = Alignment(horizontal="center")
        val_cell = ws.cell(row=stats_row + 1, column=col, value=val)
        val_cell.font = Font(bold=True, size=10, color="1D9E75")
        val_cell.alignment = Alignment(horizontal="center")
        val_cell.fill = PatternFill("solid", fgColor="E1F5EE")

    # Results table
    if summary.results:
        subject_names = [sr.subject_name for sr in summary.results[0].subject_results]
        headers = ["Pos", "Roll", "Reg. No", "Name"] + subject_names + ["GPA", "Status"]

        header_row = stats_row + 3
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.fill = PatternFill("solid", fgColor="1D9E75")
            cell.alignment = Alignment(horizontal="center")

        thin = Side(style="thin", color="CCCCCC")
        for row_idx, res in enumerate(summary.results, header_row + 1):
            fill_color = "F1EFE8" if row_idx % 2 == 0 else "FFFFFF"
            row_data = (
                [res.position or "", res.roll, res.registration_number or "", res.full_name]
                + [f"{sr.grade_letter} ({int(sr.marks)})" for sr in res.subject_results]
                + [res.gpa, res.status]
            )
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center")
                if col == len(row_data):  # Status column
                    cell.font = Font(
                        bold=True,
                        color="085041" if res.status == "PASS" else "A32D2D"
                    )

        # Freeze header
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        # Column widths
        ws.column_dimensions["A"].width = 7
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 22
        for col in range(5, 5 + len(subject_names)):
            ws.column_dimensions[get_column_letter(col)].width = 14

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
